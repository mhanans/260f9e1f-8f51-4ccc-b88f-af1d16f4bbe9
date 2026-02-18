import io
import logging
from docx import Document  # python-docx

from engine.ocr import ocr_engine

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

logger = logging.getLogger(__name__)


class FileScanner:
    IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "tif", "tiff", "bmp", "webp"}

    def extract_text(self, content: bytes, filename: str) -> str:
        """
        Extracts text from various file formats.
        Supports: .txt, .pdf, .csv, .docx, image formats.
        """
        file_ext = filename.lower().split('.')[-1]

        try:
            if file_ext == 'txt':
                return content.decode('utf-8', errors='ignore')

            if file_ext == 'csv':
                return content.decode('utf-8', errors='ignore')

            if file_ext == 'pdf':
                return self._extract_from_pdf(content)

            if file_ext == 'docx':
                return self._extract_from_docx(content)

            if file_ext in self.IMAGE_EXTENSIONS:
                return ocr_engine.extract_text_from_image(content, filename)

            return f"Unsupported file type: {file_ext}"

        except Exception as e:
            logger.error(f"Error extracting text from {filename}: {e}")
            return ""

    def _extract_from_pdf(self, content: bytes) -> str:
        chunks = self._extract_pdf_with_meta(content)
        return "\n".join(chunk.get("text", "") for chunk in chunks if chunk.get("text"))

    def _extract_from_docx(self, content: bytes) -> str:
        try:
            # Use OCR-capable extractor (includes embedded images)
            return ocr_engine.extract_text_from_docx(content)
        except Exception as e:
            logger.error(f"DOCX Extraction Error: {e}")
            return "[Error reading DOCX content]"

    def extract_with_metadata(self, content: bytes, filename: str) -> list[dict]:
        """
        Extracts text with location metadata.
        Returns: [{"text": "...", "metadata": {"page": 1, ...}}]
        """
        file_ext = filename.lower().split('.')[-1]
        try:
            if file_ext == 'pdf':
                return self._extract_pdf_with_meta(content)

            if file_ext in ['xlsx', 'xls']:
                return self._extract_excel_with_meta(content)

            if file_ext == 'docx':
                text = self._extract_from_docx(content)
                return [{"text": text, "metadata": {"type": "docx"}}]

            if file_ext in self.IMAGE_EXTENSIONS:
                text = ocr_engine.extract_text_from_image(content, filename)
                return [{"text": text, "metadata": {"type": "image", "file_ext": file_ext}}]

            # Fallback for others
            text = self.extract_text(content, filename)
            return [{"text": text, "metadata": {"type": "general"}}]
        except Exception as e:
            logger.error(f"Error extracting metadata from {filename}: {e}")
            return []

    def _extract_pdf_with_meta(self, content: bytes) -> list[dict]:
        chunks = []

        # If text extractor dependency is missing, fallback to OCR for whole PDF.
        if fitz is None:
            logger.warning("PyMuPDF (fitz) not installed, using OCR fallback for PDF.")
            text = ocr_engine.extract_text_from_pdf(content)
            if text.strip():
                return [{"text": text, "metadata": {"ocr": True, "source": "pdf_fallback"}}]
            return [{"text": "[PDF Support Missing]", "metadata": {"error": "missing_dependency"}}]

        try:
            with fitz.open(stream=content, filetype="pdf") as doc:
                for page_num, page in enumerate(doc):
                    text = page.get_text().strip()

                    if text:
                        chunks.append({
                            "text": text,
                            "metadata": {"page": page_num + 1, "ocr": False}
                        })

                    # OCR embedded images in the page.
                    image_entries = page.get_images(full=True)
                    for image_index, img in enumerate(image_entries, start=1):
                        try:
                            xref = img[0]
                            base_image = doc.extract_image(xref)
                            image_bytes = base_image.get("image", b"")
                            if not image_bytes:
                                continue
                            image_text = ocr_engine.extract_text_from_image(
                                image_bytes,
                                f"pdf_page_{page_num + 1}_img_{image_index}",
                            )
                            if image_text.strip():
                                chunks.append({
                                    "text": image_text,
                                    "metadata": {
                                        "page": page_num + 1,
                                        "ocr": True,
                                        "image_index": image_index,
                                    },
                                })
                        except Exception as inner:
                            logger.error(f"PDF image OCR error page={page_num+1} img={image_index}: {inner}")

                    # If page has no readable text and no OCR chunks yet, OCR the entire page image.
                    has_page_chunks = any(c.get("metadata", {}).get("page") == page_num + 1 for c in chunks)
                    if not has_page_chunks:
                        try:
                            pix = page.get_pixmap()
                            page_img_bytes = pix.tobytes("png")
                            page_ocr_text = ocr_engine.extract_text_from_image(
                                page_img_bytes,
                                f"pdf_page_{page_num + 1}",
                            )
                            if page_ocr_text.strip():
                                chunks.append({
                                    "text": page_ocr_text,
                                    "metadata": {"page": page_num + 1, "ocr": True, "source": "full_page"},
                                })
                        except Exception as inner:
                            logger.error(f"PDF full-page OCR error page={page_num+1}: {inner}")
        except Exception as e:
            logger.error(f"PDF Meta Extraction Error: {e}")

        return chunks

    def _extract_excel_with_meta(self, content: bytes) -> list[dict]:
        chunks = []
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    text_parts = []
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            text_parts.append(str(cell))

                    if text_parts:
                        chunks.append({
                            "text": " ".join(text_parts),
                            "metadata": {"sheet": sheet, "row": i}
                        })
        except ImportError:
            logger.error("openpyxl not installed")
            return [{"text": "Excel support missing openpyxl", "metadata": {}}]
        except Exception as e:
            logger.error(f"Excel Meta Extraction Error: {e}")
        return chunks


file_scanner = FileScanner()
